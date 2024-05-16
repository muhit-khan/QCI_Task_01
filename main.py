from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import datetime
import time
from bs4 import BeautifulSoup


# Initialize the WebDriver
driver = webdriver.Chrome()


# Get the current day of the week
today = datetime.date.today().weekday()
day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
current_day = day_names[today]

# Load your Excel file and Access the sheet for the current day
workbook = load_workbook("./Excel.xlsx")
sheet = workbook[current_day]

# Iterate through the keywords in the sheet and Search Google for the keyword
for row in range(3, sheet.max_row + 1):
    keyword = sheet.cell(row, 3).value

    driver.get("https://www.google.com")
    search_box = driver.find_element(By.NAME, "q")
    search_box.send_keys(keyword)
    
    time.sleep(2)
    
    # Declare a list to store the refined search results
    resultList = []
    
    try:
        results = driver.find_elements(By.CLASS_NAME, "wM6W7d")
        for result in results:
            # Get the outer HTML of the element
            result = str(result.get_attribute("outerHTML"))
            
            # Parse the HTML using BeautifulSoup
            soup = BeautifulSoup(result, "html.parser")
            
            # Extract the text inside the div
            div_tag = soup.find('div')
            span_tag = div_tag.find('span')
    
            # Unwrap both <div> and <span> tags
            div_tag.unwrap()
            span_tag.unwrap()
    
            if len(soup)>0:
                resultList.append(str(soup))
    
    except Exception as e:
        print(f"Error processing keyword: {keyword}. Error: {str(e)}")
    
    longest_string = max(resultList, key=len)
    shortest_string = min(resultList, key=len)

    sheet.cell(row, 4).value = longest_string
    sheet.cell(row, 5).value = shortest_string


# Save the updated Excel file and  Close the WebDriver
workbook.save("./Excel.xlsx")
driver.quit()